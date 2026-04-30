import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

SOURCE = Path("D:/Download/Copy of TRUCK ORDER MANAGEMENT.xlsx")
OUT_DIR = Path("A:/Solofleet/outputs/truck_order_bi")


def serialise(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if hasattr(value, "total_seconds"):
        return str(value)
    return value


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\n", " ").strip()
        return text or None
    return value


def text(value):
    value = clean(value)
    if value is None:
        return None
    return str(value).strip()


def number(value):
    if isinstance(value, bool) or value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def date_key(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def month_key(order):
    for field in ("load_date", "order_date"):
        value = order.get(field)
        if isinstance(value, str) and len(value) >= 7:
            return value[:7]
    return None


def is_yes(value):
    return str(value or "").strip().upper() == "YES"


def completion_status(order):
    if order["order_type"] == "Dedicated":
        return "Confirmed" if is_yes(order.get("unit_confirmed")) else "Open"
    if order.get("actual_unload_done"):
        return "Completed"
    if order.get("actual_load_done") or order.get("actual_pod_arrived"):
        return "In Transit"
    if is_yes(order.get("unit_confirmed")) or is_yes(order.get("driver_confirmed")):
        return "Confirmed"
    return "Open"


def plate_key(value):
    value = text(value)
    if not value:
        return None
    return "".join(ch for ch in value.upper() if ch.isalnum())


def top_items(counter, n=8):
    return [[label, count] for label, count in counter.most_common(n) if label]


def extract_orders(wb):
    orders = []

    ws = wb["Dedicated"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        vals = list(row)
        if row_idx > 10000:
            break
        if not any(v not in (None, "") for v in vals[:21]):
            continue
        if text(vals[0]) == "NO" or not text(vals[2]):
            continue
        orders.append(
            {
                "order_type": "Dedicated",
                "source_row": row_idx,
                "no": number(vals[0]),
                "order_date": serialise(clean(vals[1])),
                "customer": text(vals[2]),
                "plate": text(vals[3]),
                "truck_type": text(vals[4]),
                "asset": text(vals[5]),
                "temp_req": text(vals[6]),
                "load_date": serialise(clean(vals[7])),
                "load_time": serialise(clean(vals[8])),
                "origin": text(vals[11]),
                "destination": text(vals[12]),
                "sla_days": number(vals[13]),
                "unit_confirmed": text(vals[14]),
                "driver_confirmed": None,
                "driver": text(vals[15]),
                "phone": text(vals[16]),
                "vendor_driver": text(vals[18]),
                "do_no": text(vals[17]),
                "drop_point": text(vals[19]),
                "actual_load_done": None,
                "actual_pod_arrived": None,
                "actual_unload_done": None,
            }
        )

    ws = wb["On Call"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        vals = list(row)
        if row_idx > 10000:
            break
        if not any(v not in (None, "") for v in vals[:30]):
            continue
        if text(vals[1]) == "NO" or not text(vals[3]):
            continue
        orders.append(
            {
                "order_type": "On Call",
                "source_row": row_idx,
                "no": number(vals[1]),
                "order_date": serialise(clean(vals[2])),
                "customer": text(vals[3]),
                "plate": text(vals[4]),
                "truck_type": text(vals[5]),
                "asset": text(vals[6]),
                "temp_req": text(vals[7]),
                "load_date": serialise(clean(vals[12])),
                "load_time": serialise(clean(vals[13])),
                "origin": text(vals[18]),
                "destination": text(vals[19]),
                "sla_days": number(vals[20]),
                "unit_confirmed": text(vals[21]),
                "driver_confirmed": text(vals[22]),
                "driver": text(vals[23]),
                "phone": text(vals[24]),
                "vendor_driver": text(vals[25]),
                "do_no": text(vals[26]),
                "drop_point": text(vals[27]),
                "actual_load_done": serialise(clean(vals[15])),
                "actual_pod_arrived": serialise(clean(vals[16])),
                "actual_unload_done": serialise(clean(vals[17])),
            }
        )

    for order in orders:
        order["month"] = month_key(order)
        order["completion_status"] = completion_status(order)
        order["plate_key"] = plate_key(order.get("plate"))

    return orders


def extract_assets(wb):
    rows = []
    ws = wb["ASSET"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        vals = list(row)
        if row_idx > 10000:
            break
        if not text(vals[1]):
            continue
        rows.append(
            {
                "no": number(vals[0]),
                "plate": text(vals[1]),
                "asset": text(vals[2]),
                "truck_type": text(vals[3]),
                "type_truck": text(vals[4]),
                "capacity_kg": number(vals[5]),
                "capacity_cbm": number(vals[6]),
                "plate_key": plate_key(vals[1]),
            }
        )
    return rows


def extract_sla_routes(wb):
    rows = []
    ws = wb["SLA DELIVERY"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        vals = list(row)
        if row_idx > 10000:
            break
        if not text(vals[1]) or not text(vals[2]):
            continue
        if str(text(vals[1])).upper() == "ORIGINE":
            continue
        rows.append(
            {
                "no": number(vals[0]),
                "origin": text(vals[1]),
                "destination": text(vals[2]),
                "sla_days": number(vals[3]),
            }
        )
    return rows


def build_dashboard_payload(wb):
    orders = extract_orders(wb)
    assets = extract_assets(wb)
    sla_routes = extract_sla_routes(wb)

    month_counter = Counter(o["month"] for o in orders if o.get("month"))
    source_counter = Counter(o["order_type"] for o in orders)
    customer_counter = Counter(o["customer"] for o in orders if o.get("customer"))
    truck_counter = Counter(o["truck_type"] for o in orders if o.get("truck_type"))
    asset_counter = Counter(o["asset"] for o in orders if o.get("asset"))
    status_counter = Counter(o["completion_status"] for o in orders)
    used_fleet = {o["plate_key"] for o in orders if o.get("plate_key")}
    fleet_plates = {a["plate_key"] for a in assets if a.get("plate_key")}

    summary = {
        "total_orders": len(orders),
        "dedicated_orders": source_counter.get("Dedicated", 0),
        "on_call_orders": source_counter.get("On Call", 0),
        "unique_customers": len({o["customer"] for o in orders if o.get("customer")}),
        "unique_ordered_trucks": len(used_fleet),
        "fleet_units": len(fleet_plates),
        "fleet_units_used": len(used_fleet & fleet_plates),
        "completion_rate": status_counter.get("Completed", 0) / len(orders) if orders else 0,
        "confirmed_rate": (
            status_counter.get("Completed", 0)
            + status_counter.get("Confirmed", 0)
            + status_counter.get("In Transit", 0)
        )
        / len(orders)
        if orders
        else 0,
        "avg_sla_days": round(
            sum(o["sla_days"] for o in orders if isinstance(o.get("sla_days"), (int, float)))
            / max(1, len([o for o in orders if isinstance(o.get("sla_days"), (int, float))])),
            2,
        ),
        "months": sorted(month_counter.keys()),
        "month_counts": [[m, month_counter[m]] for m in sorted(month_counter.keys())],
        "source_counts": top_items(source_counter, 4),
        "top_customers": top_items(customer_counter, 10),
        "top_truck_types": top_items(truck_counter, 10),
        "asset_counts": top_items(asset_counter, 10),
        "status_counts": top_items(status_counter, 8),
    }

    return {"orders": orders, "assets": assets, "sla_routes": sla_routes, "summary": summary}


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    payload = build_dashboard_payload(wb)
    summary_lines = []
    for ws in wb.worksheets:
        summary_lines.append(f"SHEET {ws.title} rows {ws.max_row} cols {ws.max_column}")
        non_empty_seen = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx > 3000 or non_empty_seen >= 25:
                break
            if not any(v not in (None, "") for v in row):
                continue
            non_empty_seen += 1
            summary_lines.append(
                f"{idx} {json.dumps([serialise(v) for v in list(row)[:18]], ensure_ascii=False)}"
            )
        summary_lines.append("---")

    summary_lines.append("HEADER_DETAIL")
    for sheet_name, row_numbers in {
        "Dedicated": [5, 6],
        "On Call": [2, 3, 4],
        "ASSET": [2, 3],
        "SLA DELIVERY": [6, 7],
    }.items():
        ws = wb[sheet_name]
        summary_lines.append(f"SHEET {sheet_name}")
        wanted = set(row_numbers)
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx in wanted:
                summary_lines.append(
                    f"{idx} {json.dumps([serialise(v) for v in list(row)[:60]], ensure_ascii=False)}"
                )
            if idx >= max(wanted):
                break

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "source_profile.txt").write_text("\n".join(summary_lines), encoding="utf-8")
    (OUT_DIR / "dashboard_data.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "orders": len(payload["orders"]),
                "assets": len(payload["assets"]),
                "sla_routes": len(payload["sla_routes"]),
                "summary": payload["summary"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
